"""Thor (coding lieutenant) routes: /api/thor/*"""
from __future__ import annotations

import json
import logging
import subprocess
import time
import threading
from pathlib import Path

from flask import Blueprint, jsonify, request

log = logging.getLogger(__name__)
thor_bp = Blueprint("thor", __name__)

THOR_DATA = Path.home() / "thor" / "data"
EXCEL_PATH = Path.home() / "Desktop" / "brotherhood_progress.xlsx"
THOR_SCHEDULE_FILE = THOR_DATA / "wake_schedule.json"

# Track running batch process
_thor_batch_lock = threading.Lock()
_thor_batch_proc: subprocess.Popen | None = None
_thor_last_wake: float = 0.0
_thor_auto_wake_interval: int = 43200  # 12 hours in seconds


@thor_bp.route("/api/thor")
def api_thor():
    """Thor's full status."""
    try:
        # Read status file (written by Thor's reporter)
        status_file = THOR_DATA / "status.json"
        if status_file.exists():
            status = json.loads(status_file.read_text())
        else:
            status = {"state": "offline", "agent": "thor"}

        # Queue stats
        tasks_dir = THOR_DATA / "tasks"
        results_dir = THOR_DATA / "results"
        pending = completed = failed = in_progress = 0
        if tasks_dir.exists():
            for f in tasks_dir.glob("task_*.json"):
                try:
                    data = json.loads(f.read_text())
                    s = data.get("status", "")
                    if s == "pending":
                        pending += 1
                    elif s == "completed":
                        completed += 1
                    elif s == "failed":
                        failed += 1
                    elif s == "in_progress":
                        in_progress += 1
                except Exception:
                    pass

        status["queue"] = {
            "pending": pending,
            "in_progress": in_progress,
            "completed": completed,
            "failed": failed,
        }

        # Knowledge stats
        kb_index = THOR_DATA / "knowledge" / "index.json"
        if kb_index.exists():
            try:
                index = json.loads(kb_index.read_text())
                status["knowledge_entries"] = len(index)
            except Exception:
                status["knowledge_entries"] = 0
        else:
            status["knowledge_entries"] = 0

        # Brain stats from activity log
        activity_file = THOR_DATA / "activity.json"
        if activity_file.exists():
            try:
                activities = json.loads(activity_file.read_text())
                total_tokens = sum(a.get("tokens", 0) for a in activities)
                status["total_tokens"] = total_tokens
                status["total_activities"] = len(activities)
            except Exception:
                pass

        return jsonify(status)
    except Exception as e:
        return jsonify({"state": "offline", "error": str(e)})


@thor_bp.route("/api/thor/wake", methods=["POST"])
def api_thor_wake():
    """Wake Thor for a single batch run — process all pending tasks, then sleep.

    Runs Thor in batch mode as a subprocess. Returns immediately.
    """
    global _thor_batch_proc, _thor_last_wake
    with _thor_batch_lock:
        # Check if already running
        if _thor_batch_proc and _thor_batch_proc.poll() is None:
            return jsonify({"status": "already_running", "pid": _thor_batch_proc.pid})

        # Check pending tasks
        pending = 0
        tasks_dir = THOR_DATA / "tasks"
        if tasks_dir.exists():
            for f in tasks_dir.glob("task_*.json"):
                try:
                    data = json.loads(f.read_text())
                    if data.get("status") == "pending":
                        pending += 1
                except Exception:
                    pass

        if pending == 0:
            return jsonify({"status": "no_tasks", "message": "No pending tasks — Thor stays asleep"})

        # Start Thor in batch mode
        thor_home = Path.home() / "thor"
        _thor_batch_proc = subprocess.Popen(
            ["python3", "-m", "thor", "batch"],
            cwd=str(thor_home),
            stdout=open(str(THOR_DATA / "thor_daemon.log"), "a"),
            stderr=subprocess.STDOUT,
        )
        _thor_last_wake = time.time()

        # Save last wake time
        _save_schedule()

        log.info("Thor woke up (batch mode) — PID %d, %d pending tasks", _thor_batch_proc.pid, pending)
        return jsonify({
            "status": "waking",
            "pid": _thor_batch_proc.pid,
            "pending_tasks": pending,
        })


@thor_bp.route("/api/thor/wake-status")
def api_thor_wake_status():
    """Check Thor's batch status and auto-wake schedule."""
    global _thor_batch_proc, _thor_last_wake

    # Load schedule
    schedule = _load_schedule()
    auto_enabled = schedule.get("auto_enabled", True)
    interval_h = schedule.get("interval_hours", 12)
    last_wake = schedule.get("last_wake", 0)

    # Check if batch is running
    batch_running = False
    batch_pid = None
    if _thor_batch_proc:
        rc = _thor_batch_proc.poll()
        if rc is None:
            batch_running = True
            batch_pid = _thor_batch_proc.pid
        else:
            batch_pid = _thor_batch_proc.pid

    # Time until next auto-wake
    next_wake_in = 0
    if auto_enabled and last_wake > 0:
        next_wake_in = max(0, (last_wake + interval_h * 3600) - time.time())

    return jsonify({
        "batch_running": batch_running,
        "batch_pid": batch_pid,
        "auto_enabled": auto_enabled,
        "interval_hours": interval_h,
        "last_wake": last_wake,
        "last_wake_ago": f"{(time.time() - last_wake) / 3600:.1f}h" if last_wake > 0 else "never",
        "next_wake_in": f"{next_wake_in / 3600:.1f}h" if next_wake_in > 0 else "now",
    })


@thor_bp.route("/api/thor/schedule", methods=["POST"])
def api_thor_schedule():
    """Configure Thor's auto-wake schedule."""
    data = request.get_json() or {}
    schedule = _load_schedule()

    if "auto_enabled" in data:
        schedule["auto_enabled"] = bool(data["auto_enabled"])
    if "interval_hours" in data:
        schedule["interval_hours"] = max(1, min(48, int(data["interval_hours"])))

    _save_schedule(schedule)
    return jsonify({"status": "ok", "schedule": schedule})


def thor_auto_wake_check():
    """Called by dashboard refresh loop to check if Thor should auto-wake.

    Returns True if Thor was woken up.
    """
    global _thor_batch_proc, _thor_last_wake
    schedule = _load_schedule()

    if not schedule.get("auto_enabled", True):
        return False

    interval_s = schedule.get("interval_hours", 12) * 3600
    last_wake = schedule.get("last_wake", 0)

    # Check if it's time
    if time.time() - last_wake < interval_s:
        return False

    # Check if already running
    with _thor_batch_lock:
        if _thor_batch_proc and _thor_batch_proc.poll() is None:
            return False

    # Check pending tasks
    pending = 0
    tasks_dir = THOR_DATA / "tasks"
    if tasks_dir.exists():
        for f in tasks_dir.glob("task_*.json"):
            try:
                data = json.loads(f.read_text())
                if data.get("status") == "pending":
                    pending += 1
            except Exception:
                pass

    if pending == 0:
        # No tasks — update last_wake so we don't check every second
        schedule["last_wake"] = time.time()
        _save_schedule(schedule)
        return False

    # Wake Thor
    with _thor_batch_lock:
        thor_home = Path.home() / "thor"
        _thor_batch_proc = subprocess.Popen(
            ["python3", "-m", "thor", "batch"],
            cwd=str(thor_home),
            stdout=open(str(THOR_DATA / "thor_daemon.log"), "a"),
            stderr=subprocess.STDOUT,
        )
        _thor_last_wake = time.time()
        schedule["last_wake"] = _thor_last_wake
        _save_schedule(schedule)
        log.info("[AUTO-WAKE] Thor woke up — PID %d, %d pending tasks", _thor_batch_proc.pid, pending)
    return True


def _load_schedule() -> dict:
    try:
        if THOR_SCHEDULE_FILE.exists():
            return json.loads(THOR_SCHEDULE_FILE.read_text())
    except Exception:
        pass
    return {"auto_enabled": True, "interval_hours": 12, "last_wake": 0}


def _save_schedule(schedule: dict | None = None):
    global _thor_last_wake
    if schedule is None:
        schedule = _load_schedule()
    if _thor_last_wake > schedule.get("last_wake", 0):
        schedule["last_wake"] = _thor_last_wake
    try:
        THOR_SCHEDULE_FILE.parent.mkdir(parents=True, exist_ok=True)
        THOR_SCHEDULE_FILE.write_text(json.dumps(schedule, indent=2))
    except Exception:
        pass


@thor_bp.route("/api/thor/queue")
def api_thor_queue():
    """Thor's task queue."""
    try:
        tasks_dir = THOR_DATA / "tasks"
        tasks = []
        if tasks_dir.exists():
            for f in sorted(tasks_dir.glob("task_*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
                try:
                    data = json.loads(f.read_text())
                    tasks.append(data)
                except Exception:
                    pass
        return jsonify({"tasks": tasks[:50]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@thor_bp.route("/api/thor/results")
def api_thor_results():
    """Thor's recent results."""
    try:
        results_dir = THOR_DATA / "results"
        results = []
        if results_dir.exists():
            for f in sorted(results_dir.glob("result_*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
                try:
                    data = json.loads(f.read_text())
                    # Don't send full file contents over API
                    if "files_written" in data:
                        data["files_written"] = list(data["files_written"].keys())
                    results.append(data)
                except Exception:
                    pass
        return jsonify({"results": results[:20]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@thor_bp.route("/api/thor/activity")
def api_thor_activity():
    """Thor's recent activity log."""
    try:
        activity_file = THOR_DATA / "activity.json"
        if activity_file.exists():
            activities = json.loads(activity_file.read_text())
            return jsonify({"activities": activities[-30:]})
        return jsonify({"activities": []})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@thor_bp.route("/api/thor/costs")
def api_thor_costs():
    """Thor's API cost tracking report."""
    try:
        summary_file = THOR_DATA / "cost_summary.json"
        if summary_file.exists():
            summary = json.loads(summary_file.read_text())
        else:
            summary = {}

        # Calculate daily spend from cost log
        from datetime import datetime, timezone, timedelta
        from zoneinfo import ZoneInfo
        ET = ZoneInfo("America/New_York")
        today = datetime.now(ET).strftime("%Y-%m-%d")
        daily_spend = 0.0
        cost_log = THOR_DATA / "cost_log.jsonl"
        recent_calls = []
        if cost_log.exists():
            for line in cost_log.read_text().splitlines():
                if not line.strip():
                    continue
                try:
                    entry = json.loads(line)
                    if entry.get("timestamp", "").startswith(today):
                        daily_spend += entry.get("cost_usd", 0)
                    recent_calls.append(entry)
                except Exception:
                    continue

        daily_budget = 2.0
        return jsonify({
            "daily_spend_usd": round(daily_spend, 4),
            "daily_budget_usd": daily_budget,
            "daily_remaining_usd": round(daily_budget - daily_spend, 4),
            "daily_pct": round((daily_spend / daily_budget) * 100, 1) if daily_budget > 0 else 0,
            "total_calls": summary.get("total_calls", 0),
            "total_spend_usd": summary.get("total_spend_usd", 0),
            "total_input_tokens": summary.get("total_input_tokens", 0),
            "total_output_tokens": summary.get("total_output_tokens", 0),
            "by_model": summary.get("by_model", {}),
            "last_call": summary.get("last_call"),
            "recent_calls": recent_calls[-10:],
        })
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@thor_bp.route("/api/thor/review")
def api_thor_review():
    """Thor's AI code review stats."""
    try:
        from thor.core.reviewer import CodeReviewer
        cr = CodeReviewer(THOR_DATA)
        return jsonify(cr.get_stats())
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@thor_bp.route("/api/thor/codebase-index")
def api_thor_codebase_index():
    """Codebase index — functions, classes, complexity across all agents."""
    try:
        from thor.core.codebase_index import CodebaseIndex
        ci = CodebaseIndex(THOR_DATA)
        stats = ci.get_stats()
        # Per-agent summaries
        agent_summaries = {}
        for agent in (stats.get("agents_indexed") or []):
            agent_summaries[agent] = ci.get_agent_summary(agent)
        return jsonify({
            "stats": stats,
            "agents": agent_summaries,
            "stale": ci.is_stale(),
        })
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@thor_bp.route("/api/thor/codebase-index/build", methods=["POST"])
def api_thor_codebase_index_build():
    """Trigger a fresh codebase index build."""
    try:
        from thor.core.codebase_index import CodebaseIndex
        ci = CodebaseIndex(THOR_DATA)
        stats = ci.build()
        return jsonify({"status": "built", "stats": stats})
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@thor_bp.route("/api/thor/codebase-index/search")
def api_thor_codebase_search():
    """Search functions/classes in the codebase index."""
    from flask import request
    query = request.args.get("q", "")
    agent = request.args.get("agent", "")
    search_type = request.args.get("type", "function")
    if not query:
        return jsonify({"error": "Missing query parameter 'q'"}), 400
    try:
        from thor.core.codebase_index import CodebaseIndex
        ci = CodebaseIndex(THOR_DATA)
        if search_type == "class":
            results = ci.search_classes(query, agent)
        else:
            results = ci.search_functions(query, agent)
        return jsonify({"results": results, "count": len(results)})
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@thor_bp.route("/api/thor/progress")
def api_thor_progress():
    """Thor's task progress tracker — active tasks + stats."""
    try:
        from thor.core.progress import ProgressTracker
        pt = ProgressTracker(THOR_DATA)
        return jsonify({
            "active": pt.get_all_active(),
            "stats": pt.get_stats(),
        })
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@thor_bp.route("/api/thor/reflexion")
def api_thor_reflexion():
    """Thor's reflexion stats — self-debugging memory."""
    try:
        from thor.core.reflexion import ReflexionMemory
        rm = ReflexionMemory(THOR_DATA)
        return jsonify(rm.get_stats())
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@thor_bp.route("/api/thor/cache")
def api_thor_cache():
    """Thor's response cache stats."""
    try:
        from thor.core.cache import ResponseCache
        rc = ResponseCache(THOR_DATA)
        return jsonify(rc.get_stats())
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


ATLAS_DATA = Path.home() / "atlas" / "data"
SHELBY_DATA = Path.home() / "shelby" / "data"
SENTINEL_DATA = Path.home() / "sentinel" / "data"
GARVES_DATA = Path.home() / "polymarket-bot" / "data"
SOREN_DATA = Path.home() / "soren-content" / "data"
MERCURY_DATA = Path.home() / "mercury" / "data"

# Agent colors for smart action UI
AGENT_COLORS = {
    "garves": "#FFD700", "soren": "#9370DB", "shelby": "#4169E1",
    "atlas": "#32CD32", "lisa": "#FF69B4", "robotox": "#FF4500",
    "thor": "#00CED1", "dashboard": "#708090", "system": "#e0e0e0",
    "razor": "#FF1493",
}

# Project file mappings
AGENT_FILE_MAP = {
    "garves": {"root": "polymarket-bot", "key_files": ["bot/signals.py", "bot/indicators.py", "bot/main.py"]},
    "soren": {"root": "soren-content", "key_files": ["generate.py", "writer.py", "scheduler.py"]},
    "shelby": {"root": "shelby", "key_files": ["shelby.py", "app.py", "core/tools.py"]},
    "atlas": {"root": "atlas", "key_files": ["brain.py", "researcher.py", "background.py"]},
    "lisa": {"root": "mercury", "key_files": ["mercury.py", "core/brain.py", "core/publisher.py"]},
    "robotox": {"root": "sentinel", "key_files": ["sentinel.py"]},
    "thor": {"root": "thor", "key_files": ["agent.py", "core/brain.py", "core/coder.py"]},
    "dashboard": {"root": "polymarket-bot", "key_files": ["bot/live_dashboard.py", "bot/static/dashboard.js", "bot/templates/dashboard.html"]},
    "razor": {"root": "polymarket-bot", "key_files": ["razor/engine.py", "razor/scanner.py", "razor/main.py"]},
}


def _load_completed_hashes() -> set:
    """Load hashes of already-accepted/completed actions to avoid re-suggesting."""
    try:
        history_file = Path(__file__).parent.parent / "data" / "brains" / "action_history.json"
        if history_file.exists():
            data = json.loads(history_file.read_text())
            return {a.get("title_hash", "") for a in data.get("actions", [])
                    if a.get("status") in ("accepted", "completed")}
    except Exception:
        pass
    return set()


def _load_action_learnings() -> list[dict]:
    """Load learnings from past action outcomes."""
    try:
        history_file = Path(__file__).parent.parent / "data" / "brains" / "action_history.json"
        if history_file.exists():
            data = json.loads(history_file.read_text())
            return data.get("learnings", [])
    except Exception:
        pass
    return []


def _detect_installed_infrastructure() -> set[str]:
    """Detect what infrastructure/features are already built.

    Returns a set of lowercase keyword phrases. Any suggestion whose title or
    description matches multiple keywords is considered 'already built' and
    should be filtered out of smart actions.
    """
    installed = set()
    home = Path.home()
    checks = {
        # Garves features
        "regime detection":          home / "polymarket-bot" / "bot" / "regime.py",
        "fear greed":                home / "polymarket-bot" / "bot" / "regime.py",
        "sentiment correlation":     home / "polymarket-bot" / "bot" / "regime.py",
        "risk management":           home / "polymarket-bot" / "bot" / "risk.py",
        "position sizing":           home / "polymarket-bot" / "bot" / "risk.py",
        "adaptive risk":             home / "polymarket-bot" / "bot" / "risk.py",
        "conviction engine":         home / "polymarket-bot" / "bot" / "conviction.py",
        "ml predictor":              home / "polymarket-bot" / "bot" / "ml_predictor.py",
        "machine learning predictor": home / "polymarket-bot" / "bot" / "ml_predictor.py",
        "weight learner":            home / "polymarket-bot" / "bot" / "weight_learner.py",
        "straddle engine":           home / "polymarket-bot" / "bot" / "straddle.py",
        "indicator ensemble":        home / "polymarket-bot" / "bot" / "indicators.py",
        "orderbook":                 home / "polymarket-bot" / "bot" / "orderbook.py",
        "daily cycle":               home / "polymarket-bot" / "bot" / "daily_cycle.py",
        # Quant
        "backtesting":               home / "polymarket-bot" / "quant" / "main.py",
        "backtest":                  home / "polymarket-bot" / "quant" / "main.py",
        # Hawk
        "non-crypto market":         home / "polymarket-bot" / "hawk" / "main.py",
        # Atlas features
        "centralized data analytics": home / "atlas" / "brain.py",
        "data analytics module":     home / "atlas" / "brain.py",
        "knowledge base":            home / "atlas" / "brain.py",
        "research engine":           home / "atlas" / "researcher.py",
        "competitor spy":            home / "atlas" / "competitor_spy.py",
        "competitor benchmark":      home / "atlas" / "competitor_spy.py",
        # Shared intelligence
        "event bus":                 home / "shared" / "events.py",
        "feedback loop":             home / "shared" / "events.py",
        "agent memory":              home / "shared" / "agent_memory.py",
        "agent brain":               home / "shared" / "agent_brain.py",
        "llm router":                home / "shared" / "llm_client.py",
        "cost tracking":             home / "shared" / "llm_costs.jsonl",
        # Shelby features
        "task management":           home / "shelby" / "core" / "tools.py",
        "agent orchestration":       home / "shelby" / "shelby.py",
        "scheduler":                 home / "shelby" / "core" / "scheduler.py",
        # Robotox features
        "log watcher":               home / "sentinel" / "core" / "log_watcher.py",
        "health monitor":            home / "sentinel" / "sentinel.py",
        "process monitoring":        home / "sentinel" / "sentinel.py",
        "dependency checker":        home / "sentinel" / "core" / "dep_checker.py",
        # Lisa features
        "reply intelligence":        home / "mercury" / "core" / "brain.py",
        "posting scheduler":         home / "mercury" / "core" / "scheduler.py",
        # Soren features
        "a/b testing":               home / "soren-content" / "ab_testing.py",
        "ab testing":                home / "soren-content" / "ab_testing.py",
        "trend generator":           home / "soren-content" / "trend_generator.py",
        "trend analysis":            home / "soren-content" / "trend_generator.py",
        # Dashboard features
        "command center":            home / "polymarket-bot" / "bot" / "live_dashboard.py",
        "real-time dashboard":       home / "polymarket-bot" / "bot" / "live_dashboard.py",
        "analytics dashboard":       home / "polymarket-bot" / "bot" / "live_dashboard.py",
    }
    for keyword, path in checks.items():
        if path.exists():
            installed.add(keyword)
    return installed


def _is_suggestion_stale(title: str, description: str, installed: set[str]) -> bool:
    """Check if a suggestion describes something already built.

    Returns True if 2+ installed keyword phrases appear in the title+desc,
    or if any single phrase is an exact substring of the title.
    """
    text = (title + " " + description).lower()
    title_lower = title.lower()

    # Exact match in title — single hit is enough
    for keyword in installed:
        if keyword in title_lower:
            return True

    # Fuzzy: 2+ keyword hits in full text
    hits = sum(1 for keyword in installed if keyword in text)
    return hits >= 2


def _generate_smart_actions(agent_filter: str = "") -> list[dict]:
    """Generate dynamic quick actions from live agent data.

    Args:
        agent_filter: If set, only return actions for this specific agent.
    """
    actions = []
    completed_hashes = _load_completed_hashes()
    installed_infra = _detect_installed_infrastructure()

    # 1. Atlas improvements → actionable tasks
    try:
        imp_file = ATLAS_DATA / "improvements.json"
        if imp_file.exists():
            improvements = json.loads(imp_file.read_text())
            if isinstance(improvements, dict):
                for agent_name, items in improvements.items():
                    if not isinstance(items, list):
                        continue
                    for item in items[:2]:  # top 2 per agent
                        if not isinstance(item, dict):
                            continue
                        title = item.get("title", item.get("skill", item.get("suggestion", "")))[:80]
                        if not title:
                            continue
                        desc = item.get("description", item.get("detail", title))
                        # Skip suggestions for things already built
                        if _is_suggestion_stale(title, str(desc), installed_infra):
                            continue
                        files = AGENT_FILE_MAP.get(agent_name, {}).get("key_files", [])
                        actions.append({
                            "id": f"atlas_{agent_name}_{hash(title) % 10000}",
                            "title": title,
                            "description": str(desc)[:500],
                            "agent": agent_name,
                            "source": "atlas",
                            "priority": item.get("priority", "normal"),
                            "target_files": files,
                            "color": AGENT_COLORS.get(agent_name, "#888"),
                        })
    except Exception:
        pass

    # 2. Robotox bug scan → fix actions
    try:
        scan_file = SENTINEL_DATA / "scan_results.json"
        if scan_file.exists():
            scan_data = json.loads(scan_file.read_text())
            bugs = scan_data if isinstance(scan_data, list) else scan_data.get("bugs", [])
            for bug in bugs[:3]:
                if not isinstance(bug, dict):
                    continue
                agent = bug.get("agent", "system")
                title = f"Fix: {bug.get('title', bug.get('description', 'Bug'))[:60]}"
                files = bug.get("files", AGENT_FILE_MAP.get(agent, {}).get("key_files", []))
                actions.append({
                    "id": f"bug_{hash(title) % 10000}",
                    "title": title,
                    "description": bug.get("description", title)[:500],
                    "agent": agent,
                    "source": "robotox",
                    "priority": "high",
                    "target_files": files if isinstance(files, list) else [],
                    "color": AGENT_COLORS.get(agent, "#FF4500"),
                })
    except Exception:
        pass

    # 3. Garves performance → optimization actions
    try:
        trades_file = GARVES_DATA / "trades.jsonl"
        if trades_file.exists():
            trades = []
            for line in trades_file.read_text().strip().split("\n"):
                if line.strip():
                    try:
                        trades.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
            resolved = [t for t in trades if t.get("resolved")]
            if len(resolved) >= 5:
                recent = resolved[-20:]
                wins = sum(1 for t in recent if t.get("won"))
                wr = wins / len(recent) * 100 if recent else 0
                if wr < 55:
                    actions.append({
                        "id": "garves_wr_low",
                        "title": f"Garves Win Rate Low ({wr:.0f}%) — Optimize Signals",
                        "description": f"Win rate dropped to {wr:.0f}% over last {len(recent)} trades. "
                                       "Review indicator weights, thresholds, and regime detection. "
                                       "Check if market conditions changed.",
                        "agent": "garves",
                        "source": "live_data",
                        "priority": "high",
                        "target_files": ["bot/signals.py", "bot/indicators.py", "bot/regime.py"],
                        "color": AGENT_COLORS["garves"],
                    })
    except Exception:
        pass

    # 4. Soren queue health → pipeline actions
    try:
        queue_file = Path.home() / "soren-content" / "data" / "content_queue.json"
        if queue_file.exists():
            queue = json.loads(queue_file.read_text())
            items = queue if isinstance(queue, list) else queue.get("items", [])
            pending = [i for i in items if i.get("status") == "pending"]
            failed = [i for i in items if i.get("status") == "failed"]
            if len(failed) > 2:
                actions.append({
                    "id": "soren_failed_posts",
                    "title": f"Soren: {len(failed)} Failed Posts — Fix Pipeline",
                    "description": f"{len(failed)} content items failed. Check generation errors, "
                                   "API connectivity, and retry failed items.",
                    "agent": "soren",
                    "source": "live_data",
                    "priority": "high",
                    "target_files": ["soren-content/generate.py", "mercury/core/publisher.py"],
                    "color": AGENT_COLORS["soren"],
                })
            if len(pending) < 3:
                actions.append({
                    "id": "soren_low_queue",
                    "title": f"Soren Queue Low ({len(pending)} pending) — Generate Content",
                    "description": "Content queue running low. Generate a fresh batch of content "
                                   "across all pillars to keep posting schedule on track.",
                    "agent": "soren",
                    "source": "live_data",
                    "priority": "normal",
                    "target_files": ["soren-content/generate.py", "soren-content/writer.py"],
                    "color": AGENT_COLORS["soren"],
                })
    except Exception:
        pass

    # 5. Atlas KB health → maintenance actions
    try:
        kb_file = ATLAS_DATA / "knowledge_base.json"
        if kb_file.exists():
            kb = json.loads(kb_file.read_text())
            obs_count = len(kb.get("observations", []))
            learnings = kb.get("learnings", [])
            unapplied = len([l for l in learnings if not l.get("applied")])
            if obs_count > 400:
                actions.append({
                    "id": "atlas_compress_kb",
                    "title": f"Atlas KB Bloated ({obs_count} obs) — Compress",
                    "description": f"Knowledge base has {obs_count} observations (max 500). "
                                   "Run auto-summarization to compress old observations into learnings.",
                    "agent": "atlas",
                    "source": "live_data",
                    "priority": "normal",
                    "target_files": ["atlas/brain.py"],
                    "color": AGENT_COLORS["atlas"],
                })
            if unapplied > 10:
                actions.append({
                    "id": "atlas_apply_learnings",
                    "title": f"Atlas: {unapplied} Unapplied Learnings — Review",
                    "description": f"{unapplied} validated learnings haven't been applied yet. "
                                   "Review and implement the most impactful ones.",
                    "agent": "atlas",
                    "source": "live_data",
                    "priority": "normal",
                    "target_files": ["atlas/brain.py", "atlas/improvements.py"],
                    "color": AGENT_COLORS["atlas"],
                })
    except Exception:
        pass

    # 5b. Atlas learnings → actionable insights (evolving knowledge)
    try:
        kb_file = ATLAS_DATA / "knowledge_base.json"
        if kb_file.exists():
            kb = json.loads(kb_file.read_text())
            learnings = kb.get("learnings", [])
            # Only high-confidence, actionable learnings that mention specific issues
            action_keywords = ["needs", "should", "bottleneck", "declining", "failing",
                               "low", "below", "improve", "fix", "broken", "stale"]
            for learning in reversed(learnings[-50:]):
                insight = learning.get("insight", "")
                agent = learning.get("agent", "atlas")
                conf = learning.get("confidence", 0)
                if conf < 0.7 or not any(kw in insight.lower() for kw in action_keywords):
                    continue
                # Skip learnings about things already built
                if _is_suggestion_stale(insight, "", installed_infra):
                    continue
                action_id = f"learning_{agent}_{hash(insight) % 10000}"
                actions.append({
                    "id": action_id,
                    "title": insight[:80],
                    "description": f"Atlas learned (confidence: {conf:.0%}): {insight}\n\n"
                                   f"Evidence: {json.dumps(learning.get('evidence', {}), default=str)[:300]}",
                    "agent": agent,
                    "source": "atlas",
                    "priority": "high" if conf >= 0.85 else "normal",
                    "target_files": AGENT_FILE_MAP.get(agent, {}).get("key_files", []),
                    "color": AGENT_COLORS.get(agent, "#22aa44"),
                })
                if len([a for a in actions if a["id"].startswith("learning_")]) >= 5:
                    break
    except Exception:
        pass

    # 6. Shelby assessments → low-scoring agent actions
    try:
        assess_file = SHELBY_DATA / "agent_assessments.json"
        if assess_file.exists():
            assessments = json.loads(assess_file.read_text())
            if isinstance(assessments, dict):
                for agent_name, data in assessments.items():
                    if not isinstance(data, dict):
                        continue
                    score = data.get("score", data.get("rating", 100))
                    if isinstance(score, (int, float)) and score < 60:
                        actions.append({
                            "id": f"low_score_{agent_name}",
                            "title": f"{agent_name.title()} Score Low ({score}) — Investigate",
                            "description": f"Shelby rates {agent_name.title()} at {score}/100. "
                                           f"Reason: {data.get('reason', data.get('notes', 'underperforming'))}. "
                                           "Investigate and fix root cause.",
                            "agent": agent_name,
                            "source": "shelby",
                            "priority": "high",
                            "target_files": AGENT_FILE_MAP.get(agent_name, {}).get("key_files", []),
                            "color": AGENT_COLORS.get(agent_name, "#888"),
                        })
    except Exception:
        pass

    # 7. Dependency checker → security actions
    try:
        dep_file = SENTINEL_DATA / "dep_report.json"
        if dep_file.exists():
            dep_data = json.loads(dep_file.read_text())
            cves = dep_data.get("vulnerabilities", dep_data.get("cves", []))
            if cves and len(cves) > 0:
                actions.append({
                    "id": "dep_cves",
                    "title": f"{len(cves)} Security Vulnerabilities Found — Patch",
                    "description": "Robotox found vulnerable dependencies. "
                                   "Update affected packages and verify no breaking changes.",
                    "agent": "system",
                    "source": "robotox",
                    "priority": "critical",
                    "target_files": [],
                    "color": "#FF0000",
                })
    except Exception:
        pass

    # 8. Thor failed tasks → retry actions
    try:
        tasks_dir = THOR_DATA / "tasks"
        if tasks_dir.exists():
            failed_tasks = []
            for f in tasks_dir.glob("task_*.json"):
                try:
                    td = json.loads(f.read_text())
                    if td.get("status") == "failed":
                        failed_tasks.append(td)
                except Exception:
                    pass
            if failed_tasks:
                latest = sorted(failed_tasks, key=lambda x: x.get("created_at", 0), reverse=True)[0]
                actions.append({
                    "id": f"retry_{latest.get('id', 'unknown')[:8]}",
                    "title": f"Retry Failed: {latest.get('title', 'Task')[:50]}",
                    "description": f"Task failed: {latest.get('error', 'unknown error')[:200]}. "
                                   "Retry with improved context.",
                    "agent": latest.get("agent", "system"),
                    "source": "thor",
                    "priority": "normal",
                    "target_files": latest.get("target_files", []),
                    "color": AGENT_COLORS["thor"],
                })
    except Exception:
        pass

    # Filter out already-completed/accepted actions
    filtered = []
    for a in actions:
        title_hash = a.get("title", "").strip().lower()[:80]
        if title_hash not in completed_hashes:
            filtered.append(a)
    actions = filtered

    # Filter by agent if requested
    if agent_filter:
        actions = [a for a in actions if a.get("agent") == agent_filter]

    # Sort by priority (critical > high > normal > low)
    priority_order = {"critical": 0, "high": 1, "normal": 2, "low": 3}
    actions.sort(key=lambda a: priority_order.get(a.get("priority", "normal"), 2))

    # Deduplicate by id
    seen = set()
    unique = []
    for a in actions:
        if a["id"] not in seen:
            seen.add(a["id"])
            unique.append(a)

    return unique[:12]


@thor_bp.route("/api/thor/smart-actions")
def api_thor_smart_actions():
    """Generate dynamic quick actions from live agent data."""
    try:
        agent = request.args.get("agent", "")
        actions = _generate_smart_actions(agent_filter=agent)
        learnings = _load_action_learnings()[-5:]
        return jsonify({
            "actions": actions,
            "count": len(actions),
            "generated_at": time.time(),
            "recent_learnings": learnings,
        })
    except Exception as e:
        return jsonify({"actions": [], "error": str(e)})


@thor_bp.route("/api/thor/quick-action", methods=["POST"])
def api_thor_quick_action():
    """Submit a smart action or custom action as a Thor task."""
    try:
        data = request.get_json() or {}

        # Accept either a pre-built action object or action fields directly
        title = data.get("title", "")
        description = data.get("description", "")
        if not title or not description:
            return jsonify({"error": "title and description required"}), 400

        from thor.core.task_queue import CodingTask, TaskQueue
        from thor.config import ThorConfig
        cfg = ThorConfig()
        queue = TaskQueue(cfg.tasks_dir, cfg.results_dir)

        task = CodingTask(
            title=title,
            description=description,
            target_files=data.get("target_files", []),
            context_files=data.get("context_files", []),
            agent=data.get("agent", ""),
            priority=data.get("priority", "normal"),
            assigned_by=f"dashboard-smart-action:{data.get('source', 'manual')}",
        )
        task_id = queue.submit(task)
        return jsonify({"task_id": task_id, "status": "submitted", "title": title})
    except Exception as e:
        return jsonify({"error": str(e)}), 500, 500


@thor_bp.route("/api/thor/submit", methods=["POST"])
def api_thor_submit():
    """Submit a coding task to Thor via API."""
    try:
        data = request.get_json() or {}
        title = data.get("title", "")
        description = data.get("description", "")
        if not title or not description:
            return jsonify({"error": "title and description required"}), 400

        from thor.core.task_queue import CodingTask, TaskQueue
        from thor.config import ThorConfig
        cfg = ThorConfig()
        queue = TaskQueue(cfg.tasks_dir, cfg.results_dir)

        task = CodingTask(
            title=title,
            description=description,
            target_files=data.get("target_files", []),
            context_files=data.get("context_files", []),
            agent=data.get("agent", ""),
            priority=data.get("priority", "normal"),
            test_command=data.get("test_command", ""),
            assigned_by=data.get("assigned_by", "dashboard"),
        )
        task_id = queue.submit(task)
        return jsonify({"task_id": task_id, "status": "submitted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500, 500


def _apply_progress_row_style(ws, row_num, agent):
    """Apply consistent styling to a progress sheet data row."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    THIN_BORDER = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    AGENT_TEXT_COLORS = {
        "Garves": "B8860B", "Soren": "7B2D8E", "Shelby": "2B4C9B",
        "Atlas": "1E8C1E", "Lisa": "C71585", "Robotox": "CC3700",
        "Thor": "008B8B", "Dashboard": "4A4A4A", "System": "555555",
    }
    fill = (PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            if row_num % 2 == 0
            else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))

    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=(col == 7),
            horizontal="center" if col in (1, 3, 5, 8) else "left",
        )
        if col == 4:
            color = AGENT_TEXT_COLORS.get(agent, "333333")
            cell.font = Font(name="Calibri", size=10, bold=True, color=color)
        else:
            cell.font = Font(name="Calibri", size=10)


# Type mapping for Thor results
_THOR_TYPE_MAP = {
    "new feature": "Feature", "feature": "Feature", "bug fix": "Fix",
    "fix": "Fix", "improvement": "Upgrade", "upgrade": "Upgrade",
    "integration": "Integration",
}


@thor_bp.route("/api/thor/update-sheet", methods=["POST"])
def api_thor_update_sheet():
    """Scan recent Thor results and backfill any missing entries into the Excel progress sheet.

    New 8-column layout: # | Date | Time | Agent | Type | Change | Description | Status
    """
    try:
        import openpyxl
        import subprocess
        import re
        from datetime import datetime

        data_copy = THOR_DATA / "brotherhood_progress.xlsx"

        if not data_copy.exists():
            return jsonify({"error": "Excel sheet not found — run: cp ~/Desktop/brotherhood_progress.xlsx ~/thor/data/"}), 404

        wb = openpyxl.load_workbook(str(data_copy))
        ws = wb.active

        # Collect existing Change titles (col 6) to avoid duplicates
        existing = set()
        for r in range(2, ws.max_row + 1):
            change = ws.cell(r, 6).value or ""
            existing.add(change.strip())

        # Scan Thor completed results
        results_dir = THOR_DATA / "results"
        tasks_dir = THOR_DATA / "tasks"
        added = 0

        for rfile in sorted(results_dir.glob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True):
            try:
                result = json.loads(rfile.read_text())
                task_id = result.get("task_id", "")
                task_files = list(tasks_dir.glob(f"{task_id}*.json"))
                if not task_files:
                    continue
                task = json.loads(task_files[0].read_text())
                if task.get("status") != "completed":
                    continue

                change_title = task.get("title", "")[:50]
                if change_title.strip() in existing:
                    continue

                agent = (task.get("agent") or "Thor").title()
                # Clean description: no code, no markdown, max 200 chars
                desc = result.get("summary") or task.get("description", "")
                desc = re.sub(r'#+\s+', '', desc)
                desc = re.sub(r'\*{1,2}([^*]+)\*{1,2}', r'\1', desc)
                desc = re.sub(r'```[\s\S]*?```', '', desc)
                desc = desc.replace('`', '')
                desc = re.sub(r'\s+', ' ', desc).strip()[:200]

                # Detect type
                category = (task.get("category") or "improvement").lower()
                change_type = _THOR_TYPE_MAP.get(category, "Upgrade")

                created = task.get("created_at", time.time())
                dt = datetime.fromtimestamp(created)

                row_num = ws.max_row + 1
                seq = row_num - 1

                ws.cell(row=row_num, column=1, value=seq)
                ws.cell(row=row_num, column=2, value=dt.strftime("%b %d, %Y"))
                ws.cell(row=row_num, column=3, value=dt.strftime("%-I:%M %p"))
                ws.cell(row=row_num, column=4, value=agent)
                ws.cell(row=row_num, column=5, value=change_type)
                ws.cell(row=row_num, column=6, value=change_title)
                ws.cell(row=row_num, column=7, value=desc)
                ws.cell(row=row_num, column=8, value="Done")

                _apply_progress_row_style(ws, row_num, agent)
                ws.auto_filter.ref = f"A1:H{row_num}"

                existing.add(change_title.strip())
                added += 1
            except Exception:
                continue

        wb.save(str(data_copy))

        # Copy to Desktop
        try:
            subprocess.run(["cp", str(data_copy), str(EXCEL_PATH)], timeout=5)
        except Exception:
            pass

        return jsonify({"status": "ok", "added": added, "total_rows": ws.max_row})
    except Exception as e:
        log.exception("Failed to update sheet")
        return jsonify({"error": str(e)}), 500, 500


@thor_bp.route("/api/thor/update-dashboard", methods=["POST"])
def api_thor_update_dashboard():
    """Submit a task to Thor to update the dashboard with latest agent changes."""
    try:
        from thor.core.task_queue import CodingTask, TaskQueue
        from thor.config import ThorConfig
        cfg = ThorConfig()
        queue = TaskQueue(cfg.tasks_dir, cfg.results_dir)

        task = CodingTask(
            title="Update Command Center Dashboard — Sync Latest Changes",
            description=(
                "Review all recent changes across agents and update the Command Center dashboard "
                "(~/polymarket-bot/bot/templates/dashboard.html, bot/static/dashboard.js, bot/routes/) "
                "to reflect new features, endpoints, status displays, or UI sections. "
                "Check each agent's latest capabilities and ensure the dashboard accurately shows them. "
                "Only add/update what's actually missing — don't rewrite existing working sections."
            ),
            target_files=[
                str(Path.home() / "polymarket-bot/bot/templates/dashboard.html"),
                str(Path.home() / "polymarket-bot/bot/static/dashboard.js"),
            ],
            context_files=[
                str(Path.home() / "polymarket-bot/bot/routes/__init__.py"),
            ],
            agent="dashboard",
            priority="high",
            assigned_by="jordan_dashboard_button",
        )
        task_id = queue.submit(task)
        return jsonify({"task_id": task_id, "status": "submitted", "message": "Thor will update the dashboard"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500, 500


@thor_bp.route("/api/thor/update-brotherhood", methods=["POST"])
def api_thor_update_brotherhood():
    """Submit a task to Thor to update the Brotherhood Sheet HTML."""
    try:
        from thor.core.task_queue import CodingTask, TaskQueue
        from thor.config import ThorConfig
        cfg = ThorConfig()
        queue = TaskQueue(cfg.tasks_dir, cfg.results_dir)

        task = CodingTask(
            title="Update Agent Brotherhood Sheet HTML",
            description=(
                "Update ~/Desktop/Agent_Brotherhood_Sheet.html to reflect current agent capabilities, "
                "tools, intelligence meters, and hierarchy. Ensure all 10 agents (Garves, Soren, Atlas, "
                "Lisa, Shelby, Robotox, Thor, Hawk, Viper, Quant) are listed with correct skills, "
                "colors, roles, and intelligence scores. Keep the existing HTML structure and styling."
            ),
            target_files=[
                str(Path.home() / "Desktop/Agent_Brotherhood_Sheet.html"),
            ],
            context_files=[
                str(Path.home() / "polymarket-bot/bot/routes/overview.py"),
            ],
            agent="dashboard",
            priority="high",
            assigned_by="jordan_dashboard_button",
        )
        task_id = queue.submit(task)
        return jsonify({"task_id": task_id, "status": "submitted", "message": "Thor will update the Brotherhood Sheet"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500, 500
